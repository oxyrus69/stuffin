
from openpyxl import load_workbook
wb = load_workbook(r"C:\design Project\stuffing\test-results\akumulasi (html-parser).xlsx")
ws = wb[wb.sheetnames[0]]
vals = [ws[f'{c}7'].value for c in ['C','G','K','O','S','W']]   # S01 week1
print("S01 week1:", vals, "expect [768,750,684,618,660,588]")
assert vals == [768,750,684,618,660,588]
sew = [ws[f'{c}49'].value for c in ['C','G','K','O','S','W']]
ass = [ws[f'{c}50'].value for c in ['C','G','K','O','S','W']]
print("sew row49:", sew, "| ass row50:", ass)
assert sew == [9768,10278,9042,9336,8208,3282]
assert ass == [9843,10064,10344,10117,9498,432]
print("OPENPYXL VALIDATION PASS")
